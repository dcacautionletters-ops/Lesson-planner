import streamlit as st
import pandas as pd
import difflib
import re
import io
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- CONFIGURATION ---
EXCLUDE_FACULTY = ["VISHWANATH", "SHANY", "PAVITHRA"]

def get_closest_match(name, possibilities, cutoff=0.70):
    if not name or pd.isna(name): return None
    name_clean = str(name).upper().strip()
    matches = difflib.get_close_matches(name_clean, possibilities, n=1, cutoff=cutoff)
    return matches[0] if matches else None

def apply_pro_styling_and_merge(writer, sheet_name, df_original):
    workbook = writer.book
    if sheet_name not in workbook.sheetnames: return
    worksheet = workbook[sheet_name]
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11, name='Calibri')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
    
    # --- PHYSICAL MERGE LOGIC ---
    start_row = 2
    unique_courses = df_original['Course Name'].unique()
    for course in unique_courses:
        group_size = len(df_original[df_original['Course Name'] == course])
        if group_size > 1:
            end_row = start_row + group_size - 1
            worksheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            worksheet.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
            start_row = end_row + 1
        else:
            start_row += 1

    for col in worksheet.columns:
        worksheet.column_dimensions[col[0].column_letter].width = 25

def extract_section(batch_name):
    batch_str = str(batch_name).strip()
    if len(batch_str) > 2 and batch_str[-2] == " " and batch_str[-1].isalpha():
        return batch_str[-1]
    return batch_str

def process_attendance_file(uploaded_file):
    if uploaded_file is None: return pd.DataFrame()
    try:
        df = pd.read_excel(uploaded_file, header=2)
        df_clean = pd.DataFrame({
            'Batch': df.iloc[:, 6],
            'Subject': df.iloc[:, 8],
            'Hours': pd.to_numeric(df.iloc[:, 9], errors='coerce'),
            'FacultyRaw': df.iloc[:, 16]
        }).dropna(subset=['FacultyRaw'])
        
        processed = []
        for _, row in df_clean.iterrows():
            names = re.split(r',| AND ', str(row['FacultyRaw']), flags=re.IGNORECASE)
            sec_key = extract_section(row['Batch'])
            for n in names:
                processed.append({
                    'Subject': str(row['Subject']).upper().strip(),
                    'Faculty': n.upper().strip(),
                    'Hours': row['Hours'] or 0,
                    'Sec_Key': sec_key
                })
        return pd.DataFrame(processed).groupby(['Subject', 'Faculty', 'Sec_Key'], as_index=False).max()
    except Exception: return pd.DataFrame()

# --- STREAMLIT UI ---
st.set_page_config(page_title="Universal Academic Master", layout="wide")
st.title("📑 Universal Academic Reporting Hub")

lp_col, att_col = st.columns([1, 2])
with lp_col: lp_file = st.file_uploader("1. Lesson Planner", type=['xlsx'])
with att_col: att_files = st.file_uploader("2. Attendance Reports (Single or Multiple)", type=['xlsx'], accept_multiple_files=True)

if lp_file and att_files:
    if st.button("🚀 Generate Universal Consolidated Report"):
        try:
            df_lp = pd.read_excel(lp_file, header=5)
            all_att_dfs = [process_attendance_file(f) for f in att_files]
            att_data = pd.concat(all_att_dfs, ignore_index=True) if all_att_dfs else pd.DataFrame()

            sheet_groups = {
                "II BCA": ["BCA 2025", "BCA AIML 2025", "BCA DS 2025"],
                "IV BCA": ["BCA 2024"],
                "VI BCA": ["BCA 2023"],
                "II MCA": ["MCA 2025"],
                "IV MCA": ["MCA 2024"]
            }

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                for sheet_name, keys in sheet_groups.items():
                    raw_rows = []
                    pattern = "|".join(keys)
                    batch_df = df_lp[df_lp.iloc[:, 2].astype(str).str.contains(pattern, na=False)].copy()

                    for _, lp_row in batch_df.iterrows():
                        course = str(lp_row.iloc[6]).upper().strip()
                        faculty = str(lp_row.iloc[8]).upper().strip()
                        if any(ex in faculty for ex in EXCLUDE_FACULTY): continue

                        batch_full = str(lp_row.iloc[2]).strip()
                        sec_key = extract_section(batch_full)
                        m_fac = get_closest_match(faculty, att_data['Faculty'].unique().tolist())
                        m_crs = difflib.get_close_matches(course, att_data['Subject'].unique(), n=1, cutoff=0.5)

                        actual = 0
                        if m_crs and m_fac:
                            match = att_data[(att_data['Sec_Key'] == sec_key) & (att_data['Subject'] == m_crs[0]) & (att_data['Faculty'] == m_fac)]
                            actual = match['Hours'].iloc[0] if not match.empty else 0

                        raw_rows.append({
                            'Course Name': lp_row.iloc[6],
                            'Batch': batch_full,
                            'Faculty Name': lp_row.iloc[8],
                            'Planned Sessions': pd.to_numeric(lp_row.iloc[10], errors='coerce') or 0,
                            'As per Time Table': lp_row.iloc[11],
                            'No of sessions taken': lp_row.iloc[16],
                            'Syllabus Coverage %': lp_row.iloc[18],
                            'Actual Hours Conducted': actual
                        })

                    if raw_rows:
                        df_res = pd.DataFrame(raw_rows).sort_values(by=['Course Name', 'Batch'])
                        df_res['Deviation'] = df_res['Actual Hours Conducted'] - df_res['Planned Sessions']
                        df_for_merge = df_res.copy()

                        final_formatted = []
                        last_course, sl_no = "", 1
                        for _, row in df_res.iterrows():
                            current_course = row['Course Name']
                            new_row = row.copy()
                            if current_course != last_course:
                                new_row['Sl No.'], sl_no, last_course = sl_no, sl_no+1, current_course
                            else:
                                new_row['Sl No.'], new_row['Course Name'] = "", ""
                            final_formatted.append(new_row)

                        final_df = pd.DataFrame(final_formatted)[['Sl No.', 'Course Name', 'Batch', 'Faculty Name', 'Planned Sessions', 'As per Time Table', 'No of sessions taken', 'Syllabus Coverage %', 'Actual Hours Conducted', 'Deviation']]
                        final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        apply_pro_styling_and_merge(writer, sheet_name, df_for_merge)

            st.success("✨ Universal Report Generated Successfully!")
            st.download_button("📥 Download Final Report", output.getvalue(), "Universal_Academic_Consolidated_Report.xlsx")
        except Exception as e:
            st.error(f"Error: {e}")
